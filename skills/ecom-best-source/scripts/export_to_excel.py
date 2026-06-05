#!/usr/bin/env python3
"""
比价结果Excel生成脚本
输入：JSON数据（productInfo + results）
输出：格式化的Excel比价表
"""

import json
import sys
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter


# 列定义
COLUMNS = [
    ("A", "序号", 6),
    ("B", "平台", 8),
    ("C", "商品链接", 20),
    ("D", "店铺名称", 20),
    ("E", "SKU名称", 25),
    ("F", "SKU匹配", 14),
    ("G", "起批量总价", 12),
    ("H", "运费", 10),
    ("I", "折算单价", 12),
    ("J", "京东售价", 12),
    ("K", "差价", 10),
    ("L", "差价比例", 10),
    ("M", "搜索方式", 12),
    ("N", "旺旺状态", 12),
    ("O", "旺旺对话", 40),
    ("P", "备注", 25),
]

# 样式定义
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True)
SUBTITLE_FONT = Font(name="微软雅黑", size=10, color="666666")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DATA_FONT = Font(name="微软雅黑", size=10)
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
RED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
GREEN_FONT = Font(name="微软雅黑", size=10, color="2E7D32")
ORANGE_FONT = Font(name="微软雅黑", size=10, color="E65100")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def create_comparison_excel(data, output_path):
    """生成比价结果Excel"""
    product_info = data["productInfo"]
    results = data["results"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "比价结果"

    # === 表头信息区 ===

    # 第1行：标题
    ws.merge_cells("A1:P1")
    cell = ws["A1"]
    cell.value = "供应商比价结果表"
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 第2行：商品信息
    ws.merge_cells("A2:P2")
    cell = ws["A2"]
    jd_price = product_info.get("price", "N/A")
    batch_qty = product_info.get("batchQuantity", "N/A")
    cell.value = (
        f"京东商品：{product_info.get('title', 'N/A')} ｜ "
        f"SKU：{product_info.get('skuName', 'N/A')} ｜ "
        f"京东售价：¥{jd_price} ｜ "
        f"起批量：{batch_qty}"
    )
    cell.font = SUBTITLE_FONT
    cell.alignment = LEFT_ALIGN
    ws.row_dimensions[2].height = 22

    # 第3行：生成信息
    ws.merge_cells("A3:P3")
    cell = ws["A3"]
    cell.value = (
        f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')} ｜ "
        f"搜索范围：1688+淘宝 ｜ "
        f"收货地：武汉"
    )
    cell.font = SUBTITLE_FONT
    cell.alignment = LEFT_ALIGN
    ws.row_dimensions[3].height = 22

    # 第4行空行
    ws.row_dimensions[4].height = 8

    # === 数据列标题（第5行）===
    header_row = 5
    for col_letter, col_name, col_width in COLUMNS:
        cell = ws[f"{col_letter}{header_row}"]
        cell.value = col_name
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = col_width

    ws.row_dimensions[header_row].height = 28

    # === 数据行 ===
    data_start_row = 6
    jd_price_val = float(jd_price) if isinstance(jd_price, (int, float)) else 0

    for i, item in enumerate(results):
        row = data_start_row + i

        # 序号
        ws[f"A{row}"] = i + 1
        ws[f"A{row}"].alignment = CENTER_ALIGN

        # 平台
        ws[f"B{row}"] = item.get("platform", "")
        ws[f"B{row}"].alignment = CENTER_ALIGN

        # 商品链接（超链接）
        link = item.get("link", "")
        ws[f"C{row}"].value = "点击查看"
        if link:
            ws[f"C{row}"].hyperlink = link
            ws[f"C{row}"].font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
        ws[f"C{row}"].alignment = CENTER_ALIGN

        # 店铺名称
        ws[f"D{row}"] = item.get("shopName", "")
        ws[f"D{row}"].alignment = LEFT_ALIGN

        # SKU名称
        ws[f"E{row}"] = item.get("skuName", "")
        ws[f"E{row}"].alignment = LEFT_ALIGN

        # SKU匹配等级
        match_level = item.get("skuMatchLevel", "")
        ws[f"F{row}"] = match_level
        ws[f"F{row}"].alignment = CENTER_ALIGN
        if match_level == "完全一致":
            ws[f"F{row}"].font = GREEN_FONT
        elif match_level in ("不一致", "SKU不一致"):
            ws[f"F{row}"].font = ORANGE_FONT

        # 起批量总价
        batch_price = item.get("batchPrice")
        ws[f"G{row}"] = batch_price if batch_price is not None else ""
        ws[f"G{row}"].number_format = "#,##0.00"
        ws[f"G{row}"].alignment = CENTER_ALIGN

        # 运费
        shipping = item.get("shipping")
        ws[f"H{row}"] = shipping if shipping is not None else item.get("shippingNote", "")
        if isinstance(shipping, (int, float)):
            ws[f"H{row}"].number_format = "#,##0.00"
        ws[f"H{row}"].alignment = CENTER_ALIGN

        # 折算单价
        unit_price = item.get("unitPrice")
        ws[f"I{row}"] = unit_price if unit_price is not None else ""
        ws[f"I{row}"].number_format = "#,##0.00"
        ws[f"I{row}"].alignment = CENTER_ALIGN

        # 京东售价
        ws[f"J{row}"] = jd_price_val
        ws[f"J{row}"].number_format = "#,##0.00"
        ws[f"J{row}"].alignment = CENTER_ALIGN

        # 差价
        price_diff = item.get("priceDiff")
        ws[f"K{row}"] = price_diff if price_diff is not None else ""
        ws[f"K{row}"].number_format = "#,##0.00"
        ws[f"K{row}"].alignment = CENTER_ALIGN

        # 差价比例
        diff_pct = item.get("priceDiffPercent", "")
        ws[f"L{row}"] = diff_pct
        ws[f"L{row}"].alignment = CENTER_ALIGN

        # 搜索方式
        ws[f"M{row}"] = item.get("searchType", "")
        ws[f"M{row}"].alignment = CENTER_ALIGN

        # 旺旺状态
        ws[f"N{row}"] = item.get("wangwangStatus", "")
        ws[f"N{row}"].alignment = CENTER_ALIGN

        # 旺旺对话
        ws[f"O{row}"] = item.get("wangwangMessage", "")
        ws[f"O{row}"].alignment = LEFT_ALIGN

        # 备注
        ws[f"P{row}"] = item.get("note", "")
        ws[f"P{row}"].alignment = LEFT_ALIGN

        # 行级样式
        for col_letter, _, _ in COLUMNS:
            cell = ws[f"{col_letter}{row}"]
            cell.border = THIN_BORDER
            if not cell.font or cell.font == Font():
                cell.font = DATA_FONT

        # 条件格式：差价>0绿色，<0红色
        if isinstance(price_diff, (int, float)):
            row_fill = GREEN_FILL if price_diff > 0 else RED_FILL if price_diff < 0 else None
            if row_fill:
                for col_letter, _, _ in COLUMNS:
                    ws[f"{col_letter}{row}"].fill = row_fill
        elif i % 2 == 1:
            # 交替行色（仅在无条件格式时）
            for col_letter, _, _ in COLUMNS:
                ws[f"{col_letter}{row}"].fill = ALT_ROW_FILL

        ws.row_dimensions[row].height = 24

    # === 汇总行 ===
    last_data_row = data_start_row + len(results)
    summary_row = last_data_row + 1

    cheaper_count = sum(1 for r in results if isinstance(r.get("priceDiff"), (int, float)) and r["priceDiff"] > 0)
    exact_match = sum(1 for r in results if r.get("skuMatchLevel") == "完全一致")
    ww_sent = sum(1 for r in results if r.get("wangwangStatus") == "已发送")

    ws.merge_cells(f"A{summary_row}:P{summary_row}")
    summary_cell = ws[f"A{summary_row}"]
    summary_cell.value = (
        f"总计搜索到：{len(results)}个供应商 ｜ "
        f"比京东便宜的：{cheaper_count}个 ｜ "
        f"SKU完全一致的：{exact_match}个 ｜ "
        f"旺旺已发送：{ww_sent}个"
    )
    summary_cell.font = Font(name="微软雅黑", size=10, bold=True, color="333333")
    summary_cell.alignment = LEFT_ALIGN

    # === 冻结窗格和筛选 ===
    ws.freeze_panes = f"A{data_start_row}"
    ws.auto_filter.ref = f"A{header_row}:P{last_data_row - 1}"

    # 保存
    wb.save(output_path)
    print(f"Excel saved: {output_path}")
    return output_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description="生成比价结果Excel")
    parser.add_argument("--input", required=True, help="JSON数据文件路径或JSON字符串")
    parser.add_argument("--output", required=True, help="输出Excel文件路径")
    args = parser.parse_args()

    # 读取输入数据
    if os.path.isfile(args.input):
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = json.loads(args.input)

    create_comparison_excel(data, args.output)


if __name__ == "__main__":
    main()
